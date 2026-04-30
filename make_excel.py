"""
backtest_30m_summary.xlsx 생성
Usage: python make_excel.py
"""
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "백테스트_30m_요약.xlsx"

BG_DARK   = "0D1117"
BG_HEADER = "161B22"
BG_ROW1   = "1C2128"
BG_ROW2   = "21262D"
FG_WHITE  = "E6EDF3"
FG_GRAY   = "8B949E"
FG_GREEN  = "3FB950"
FG_RED    = "F85149"
FG_BLUE   = "58A6FF"
FG_YELLOW = "D29922"

def hfill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=10): return Font(color=hex_, bold=bold, size=sz, name="Calibri")
def align(h="center", v="center"): return Alignment(horizontal=h, vertical=v, wrap_text=True)
def thin_border():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header_row(ws, row, cols):
    for c, (val, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = hfill(BG_HEADER)
        cell.font      = font(FG_BLUE, bold=True, sz=9)
        cell.alignment = align()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = width

def write_data_row(ws, row, vals, fg=None, bg=None, bold=False):
    bg = bg or (BG_ROW1 if row % 2 == 0 else BG_ROW2)
    for c, val in enumerate(vals, 1):
        f  = fg[c-1] if (fg and c-1 < len(fg)) else FG_WHITE
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = hfill(bg)
        cell.font      = font(f, bold=bold, sz=9)
        cell.alignment = align()
        cell.border    = thin_border()

def set_ws_style(ws):
    ws.sheet_view.showGridLines = False

# Sheet 1: Strategy Overview
def sheet_overview(wb):
    ws = wb.create_sheet("전략 개요")
    set_ws_style(ws)

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = "캔들 반전 전략 - BTCUSDT 30분봉"
    t.fill      = hfill(BG_DARK)
    t.font      = font(FG_BLUE, bold=True, sz=13)
    t.alignment = align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value     = "장대봉 + 반대봉 역추세 전략 | Gate.io 선물"
    s.fill      = hfill(BG_DARK)
    s.font      = font(FG_GRAY, sz=9)
    s.alignment = align()

    ws.merge_cells("A3:D3")
    ph = ws["A3"]
    ph.value     = "파라미터"
    ph.fill      = hfill(BG_HEADER)
    ph.font      = font(FG_YELLOW, bold=True, sz=10)
    ph.alignment = align()

    write_header_row(ws, 4, [("파라미터",22),("값",18),("설명",40),("비고",18)])
    params = [
        ("big_mult",  "2.4",   "신호봉 몸통 >= 평균 몸통 x 배수", "OOS 최적화"),
        ("cover_pct", "0.7",   "반전봉 겹침 비율 기준",            "OOS 최적화"),
        ("rr_ratio",  "4.3",   "손익비 (RR)",                      "OOS 최적화"),
        ("avg_len",   "13",    "평균 몸통 롤링 윈도우",             "OOS 최적화"),
        ("taker_fee", "0.01%", "Gate.io 표준 0.05% 테이커",        "내성 한계 0.11%"),
    ]
    for i, row in enumerate(params, 5):
        write_data_row(ws, i, row, fg=[FG_BLUE, FG_GREEN, FG_GRAY, FG_GRAY])

    ws.merge_cells("A11:D11")
    pw = ws["A11"]
    pw.value     = "백테스트 기간"
    pw.fill      = hfill(BG_HEADER)
    pw.font      = font(FG_YELLOW, bold=True, sz=10)
    pw.alignment = align()

    write_header_row(ws, 12, [("구간",22),("기간",18),("목적",40),("비고",18)])
    period_data = [
        ("IS  (인샘플)",    "2020-01 ~ 2022-12", "파라미터 최적화 (그리드서치)", "기준"),
        ("OOS (아웃샘플)",  "2023-01 ~ 2026-04", "실전 성과 검증",               "핵심 평가"),
    ]
    for i, row in enumerate(period_data, 13):
        write_data_row(ws, i, row)

    ws.merge_cells("A16:D16")
    perf = ws["A16"]
    perf.value     = "OOS 성과 요약"
    perf.fill      = hfill(BG_HEADER)
    perf.font      = font(FG_YELLOW, bold=True, sz=10)
    perf.alignment = align()

    write_header_row(ws, 17, [("지표",22),("IS 값",18),("OOS 값",40),("목표",18)])
    perf_data = [
        ("승률",              "~30% 추정",   "26.5%",          ">25%"),
        ("기대값",            "+0.2R 추정",  "+0.192R",        ">0"),
        ("MDD (R)",           "~9R 추정",    "17.4R",          "<20R"),
        ("최대 연속 손실",    "~10 추정",    "17연패 (실측)",   "<20"),
        ("풀 켈리",           "--",           "17.6%",          "--"),
        ("하프 켈리",         "--",           "8.8%",           "--"),
        ("권장 리스크",       "--",           "3.9%",           "복리 MDD <50%"),
        ("월 복리 수익 (3.9%)", "--",         "~+10%",          "추정"),
    ]
    for i, row in enumerate(perf_data, 18):
        is_rec = row[0] == "권장 리스크"
        bg = "1A3320" if is_rec else None
        write_data_row(ws, i, row, fg=[FG_WHITE, FG_WHITE, FG_GREEN, FG_GRAY], bg=bg)


# Sheet 2: Kelly & Risk
def sheet_kelly(wb):
    ws = wb.create_sheet("Kelly & 리스크")
    set_ws_style(ws)

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "켈리 기준 & 리스크 관리"
    t.fill      = hfill(BG_DARK)
    t.font      = font(FG_BLUE, bold=True, sz=13)
    t.alignment = align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    ws["A2"].value     = "f* = (WR x RR - (1-WR)) / RR  |  OOS: WR=26.5%  RR=4.3  MDD=17.4R"
    ws["A2"].fill      = hfill(BG_DARK)
    ws["A2"].font      = font(FG_GRAY, sz=9)
    ws["A2"].alignment = align()

    write_header_row(ws, 3, [
        ("리스크 %", 12), ("단순 MDD %", 16), ("복리 MDD %", 18),
        ("월 복리 수익", 18), ("판정", 26)
    ])

    wr  = 0.265; rr = 4.3; mdd_r = 17.4
    full_k = (wr * rr - (1 - wr)) / rr * 100
    half_k = full_k / 2

    risks = [0.3, 0.5, 0.75, 1.0, 1.3, 2.0, 3.0, 3.9, round(half_k,2), round(full_k,2)]
    seen = set()
    rows_data = []
    for p in risks:
        p = round(p, 2)
        if p in seen or p <= 0:
            continue
        seen.add(p)
        simple   = mdd_r * p
        compound = (1 - (1 - p/100)**mdd_r) * 100
        monthly  = ((1 + 0.192 * p / 100) ** 30 - 1) * 100
        if compound < 20:    verdict = "[OK] 매우 안전"
        elif compound < 35:  verdict = "[OK] 안전"
        elif compound < 50:  verdict = "[!]  주의"
        elif compound < 70:  verdict = "[X]  위험"
        else:                verdict = "[!!] 파탄 위험"
        rows_data.append((p, simple, compound, monthly, verdict))

    for i, (p, s, c, m, v) in enumerate(rows_data, 4):
        is_opt = abs(p - 3.9) < 0.05
        is_hk  = abs(p - half_k) < 0.15
        is_fk  = abs(p - full_k) < 0.15
        if c < 35:    c_color = FG_GREEN
        elif c < 50:  c_color = FG_YELLOW
        else:         c_color = FG_RED
        bg = "1A3320" if is_opt else None
        tag = ""
        if is_opt:  tag += " <-- 권장"
        if is_hk and not is_opt: tag += " <-- 하프 켈리"
        if is_fk and not is_opt: tag += " <-- 풀 켈리"
        write_data_row(ws, i, [
            f"{p:.2f}%", f"{s:.1f}%", f"{c:.1f}%", f"+{m:.1f}%", v + tag
        ], fg=[FG_WHITE, FG_WHITE, c_color, FG_GREEN, FG_WHITE], bg=bg)
        if is_opt:
            for col in range(1, 6):
                ws.cell(row=i, column=col).font = font(
                    FG_GREEN if col != 3 else c_color, bold=True, sz=9)

    summary_row = len(rows_data) + 5
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    cell = ws[f"A{summary_row}"]
    cell.value     = (f"권장: 리스크 3.9% | 복리 MDD ~50% | 월 수익 ~+10% | "
                      f"풀 켈리={full_k:.1f}%  하프 켈리={half_k:.1f}%")
    cell.fill      = hfill("1A3320")
    cell.font      = font(FG_GREEN, bold=True, sz=10)
    cell.alignment = align()


# Sheet 3: Consecutive Losses
def sheet_consec(wb):
    ws = wb.create_sheet("연속 손실 구간")
    set_ws_style(ws)

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = "17연패 구간 (OOS 실측 확인)"
    t.fill      = hfill(BG_DARK)
    t.font      = font(FG_RED, bold=True, sz=13)
    t.alignment = align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value     = "기간: 2024-03-08 ~ 2024-05-14 (67일) | MDD 17.4R 실측"
    ws["A2"].fill      = hfill(BG_DARK)
    ws["A2"].font      = font(FG_GRAY, sz=9)
    ws["A2"].alignment = align()

    write_header_row(ws, 3, [("지표",26),("값",20),("설명",36),("영향",20)])
    data = [
        ("최대 연속 손실",      "17연패",      "2024-03-08 ~ 2024-05-14",  "67 거래일"),
        ("MDD (R 단위)",        "17.4R",       "1R = 리스크 금액 1배",      "최대 낙폭"),
        ("MDD @ 리스크 1.0%",  "17.4%",       "계좌 낙폭",                 "회복 가능"),
        ("MDD @ 리스크 3.9%",  "~50%",        "복리 MDD 공식 적용",        "회복 어려움"),
        ("BTC 가격 구간",       "$62k-$71k",   "2024년 3~5월",              "횡보/상승 구간"),
        ("신호 충돌",           "역추세 전략", "추세 상승 중 역추세 진입",  "손실 원인"),
        ("필요 회복 수익률",    "+100%",       "-50% 낙폭 기준",            "리스크 3.9% 시"),
    ]
    for i, row in enumerate(data, 4):
        write_data_row(ws, i, row, fg=[FG_WHITE, FG_RED, FG_GRAY, FG_YELLOW])

    ws.merge_cells("A12:D12")
    ws["A12"].value     = "리스크 완화 방안"
    ws["A12"].fill      = hfill(BG_HEADER)
    ws["A12"].font      = font(FG_YELLOW, bold=True, sz=10)
    ws["A12"].alignment = align()

    write_header_row(ws, 13, [("방안",26),("실행",20),("효과",36),("분류",20)])
    mitigations = [
        ("보수적 시작",    "1.3% 리스크로 시작",      "복리 MDD ~20%",         "점진적 증액"),
        ("하드 스탑 룰",  "계좌 -30% 시 일시 중지",  "강제 점검 후 재개",      "서킷 브레이커"),
        ("켈리 비율",     "3.9% 권장 (하프 켈리)",   "수익과 파탄 위험 균형",  "권장"),
        ("추세 필터",     "강한 추세 시 신호 스킵",   "역추세 오신호 감소",     "전략 개선"),
    ]
    for i, row in enumerate(mitigations, 14):
        write_data_row(ws, i, row)


# Sheet 4: Fee Tolerance
def sheet_fees(wb):
    ws = wb.create_sheet("수수료 내성")
    set_ws_style(ws)

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = "수수료 내성 분석"
    t.fill      = hfill(BG_DARK)
    t.font      = font(FG_BLUE, bold=True, sz=13)
    t.alignment = align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value     = "30분봉 평균 자연 레버리지 ~175x | 실효 수수료 = 테이커 수수료 x 레버리지 x 양방향"
    ws["A2"].fill      = hfill(BG_DARK)
    ws["A2"].font      = font(FG_GRAY, sz=9)
    ws["A2"].alignment = align()

    write_header_row(ws, 3, [("테이커 수수료",16),("실효 수수료/거래 (R)",22),("OOS 기대값",22),("판정",24)])

    base_exp = 0.386
    avg_lev  = 175.0
    fees = [0.00, 0.01, 0.02, 0.05, 0.075, 0.10, 0.15, 0.20]
    for i, fee_pct in enumerate(fees, 4):
        fee_dec    = fee_pct / 100
        eff_fee    = fee_dec * avg_lev * 2
        expectancy = base_exp - eff_fee
        if expectancy > 0.2:    verdict = "[OK] 우수"
        elif expectancy > 0:    verdict = "[OK] 수익"
        elif expectancy > -0.1: verdict = "[!]  경계"
        else:                   verdict = "[X]  비수익"
        color = FG_GREEN if expectancy > 0 else FG_RED
        is_gate = abs(fee_pct - 0.05) < 0.001
        bg = "1A3320" if is_gate else None
        write_data_row(ws, i, [
            f"{fee_pct:.3f}%",
            f"{eff_fee:.3f}R",
            f"{expectancy:+.3f}R",
            verdict + (" <-- Gate.io 기본" if is_gate else ""),
        ], fg=[FG_WHITE, FG_WHITE, color, FG_WHITE], bg=bg)

    summary_row = len(fees) + 5
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    cell = ws[f"A{summary_row}"]
    cell.value     = "Gate.io 테이커 0.05% | OOS 기대값: +0.234R | 손익분기 수수료: ~0.11%"
    cell.fill      = hfill("1A3320")
    cell.font      = font(FG_GREEN, bold=True, sz=10)
    cell.alignment = align()

    exc_row = summary_row + 2
    ws.merge_cells(f"A{exc_row}:D{exc_row}")
    ws[f"A{exc_row}"].value     = "거래소별 수수료 비교"
    ws[f"A{exc_row}"].fill      = hfill(BG_HEADER)
    ws[f"A{exc_row}"].font      = font(FG_YELLOW, bold=True, sz=10)
    ws[f"A{exc_row}"].alignment = align()

    write_header_row(ws, exc_row+1, [("거래소",16),("테이커 수수료",22),("상태",22),("비고",24)])
    exchanges = [
        ("Gate.io",  "0.050%", "최적",    "기본 선택"),
        ("Binance",  "0.040%", "더 좋음", "높은 거래량 필요"),
        ("Bybit",    "0.055%", "OK",       "약간 높음"),
        ("OKX",      "0.050%", "OK",       "Gate.io 동일"),
        ("Bitget",   "0.060%", "OK",       "허용 범위 내"),
        ("KuCoin",   "0.060%", "OK",       "허용 범위 내"),
        ("Kraken",   "0.100%", "경계",     "손익분기 한계"),
    ]
    for i, row in enumerate(exchanges, exc_row+2):
        fee_val = float(row[1].replace("%",""))
        fg = [FG_WHITE,
              FG_GREEN if fee_val <= 0.05 else FG_YELLOW,
              FG_GREEN if fee_val <= 0.06 else FG_YELLOW,
              FG_GRAY]
        write_data_row(ws, i, row, fg=fg)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_overview(wb)
    sheet_kelly(wb)
    sheet_consec(wb)
    sheet_fees(wb)

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
