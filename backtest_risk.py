"""
30분봉·1시간봉 1등 파라미터 — 리스크 0.5~5% 복리 월별 성적 엑셀
"""
import io, zipfile, warnings
from pathlib import Path
import numpy as np
import pandas as pd
import requests
from numba import njit
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

CACHE_DIR = Path("cache")
SYMBOL    = "BTCUSDT"
INIT_CAP  = 10_000   # 초기 자본 달러 (비율 계산용 기준)

PARAMS = {
    "30m": dict(big_mult=2.4, cover_pct=0.7, rr_ratio=4.3, avg_len=13, taker_fee=0.0001),
    "1h":  dict(big_mult=2.7, cover_pct=0.5, rr_ratio=4.7, avg_len=22, taker_fee=0.0001),
}
RISK_PCT_LIST = [round(i * 0.005, 4) for i in range(1, 11)]  # 0.5%~5%, 0.5%씩

IS_START  = pd.Timestamp("2020-01-01", tz="UTC")
IS_END    = pd.Timestamp("2023-01-01", tz="UTC")
OOS_START = pd.Timestamp("2023-01-01", tz="UTC")
OOS_END   = pd.Timestamp("2026-04-01", tz="UTC")
RESAMPLE_MAP = {"30m": "30min", "1h": "1h"}

# ── 색상 ─────────────────────────────────────────────────────────────
C_DARK   = "1F3864"
C_WIN    = "C6EFCE"
C_LOSS   = "FFC7CE"
C_IS_HDR = "375623"
C_OOS_HDR= "833C00"
C_IS_BG  = "EBF5D5"
C_OOS_BG = "FFF2CC"
C_WHITE  = "FFFFFF"
C_GRAY   = "F2F2F2"
C_BORDER = "BFBFBF"
TF_COLOR = {"30m": "70AD47", "1h": "FF5050"}


def _fill(h): return PatternFill(fill_type="solid", fgColor=h)
def _font(bold=False, size=9, color="000000"):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color)
def _center(): return Alignment(horizontal="center", vertical="center")
def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


# ── 데이터 다운로드 ───────────────────────────────────────────────────
_BASE = "https://data.binance.vision/data/spot/monthly/klines"

def _fetch_month(symbol, year, month):
    fname = f"{symbol}-1m-{year}-{month:02d}"
    url   = f"{_BASE}/{symbol}/1m/{fname}.zip"
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        with z.open(z.namelist()[0]) as f:
            df = pd.read_csv(f, header=None,
                             names=["ts","open","high","low","close","volume",
                                    "close_ts","qvol","ntrades","tb","tq","ign"],
                             usecols=["ts","open","high","low","close","volume"])
    for c in ["ts","open","high","low","close","volume"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna()
    ts_vals = df["ts"].astype("int64").values
    unit = "us" if ts_vals.max() > 1e13 else "ms"
    df["ts"] = pd.to_datetime(ts_vals, unit=unit, utc=True)
    return df

def load_1m():
    cache = CACHE_DIR / f"{SYMBOL}_1m_2020_2026.parquet"
    if cache.exists():
        print(f"캐시 로드: {cache}", flush=True)
        return pd.read_parquet(cache)
    cutoff = pd.Timestamp("2026-04-01")
    months = [(y, m) for y in range(2020, 2027)
              for m in range(1, 13)
              if pd.Timestamp(year=y, month=m, day=1) < cutoff]
    print(f"Binance 다운로드: {SYMBOL} 1분봉 ({len(months)}개월)", flush=True)
    frames = []
    for i, (y, m) in enumerate(months, 1):
        print(f"  [{i}/{len(months)}] {y}-{m:02d}...", end=" ", flush=True)
        try:
            df_m = _fetch_month(SYMBOL, y, m)
            frames.append(df_m)
            print(f"{len(df_m):,}봉", flush=True)
        except Exception as e:
            print(f"건너뜀 ({e})", flush=True)
    df = pd.concat(frames, ignore_index=True)
    df = df.drop_duplicates("ts").sort_values("ts").reset_index(drop=True)
    df.to_parquet(cache)
    return df

def resample(df_1m, tf):
    rule = RESAMPLE_MAP[tf]
    df = df_1m.set_index("ts")
    return df.resample(rule, label="left", closed="left").agg(
        open=("open","first"), high=("high","max"),
        low=("low","min"), close=("close","last"), volume=("volume","sum")
    ).dropna(subset=["open"]).reset_index()


# ── 신호 감지 ─────────────────────────────────────────────────────────
def _rolling_mean(arr, window):
    result = np.full(len(arr), np.nan)
    if len(arr) < window:
        return result
    cs = np.cumsum(arr)
    result[window - 1] = cs[window - 1] / window
    if len(arr) > window:
        result[window:] = (cs[window:] - cs[:len(arr) - window]) / window
    return result

def detect_signals(df, big_mult, cover_pct, avg_len):
    close = df["close"].values
    open_ = df["open"].values
    high  = df["high"].values
    low   = df["low"].values
    n     = len(df)
    body  = np.abs(close - open_)
    avg_body = _rolling_mean(body, int(avg_len))
    is_big  = body >= avg_body * big_mult
    is_bull = close > open_
    is_bear = close < open_
    prev_top = np.maximum(open_[:-1], close[:-1])
    prev_bot = np.minimum(open_[:-1], close[:-1])
    curr_top = np.maximum(open_[1:],  close[1:])
    curr_bot = np.minimum(open_[1:],  close[1:])
    prev_body = body[:-1]
    overlap  = np.minimum(prev_top, curr_top) - np.maximum(prev_bot, curr_bot)
    cover    = np.where(prev_body > 0, overlap / prev_body, 0.0)
    short_sig = np.zeros(n, dtype=bool)
    long_sig  = np.zeros(n, dtype=bool)
    short_sig[1:] = is_big[:-1] & is_bull[:-1] & is_bear[1:] & (cover >= cover_pct)
    long_sig[1:]  = is_big[:-1] & is_bear[:-1] & is_bull[1:] & (cover >= cover_pct)
    return short_sig, long_sig


# ── Numba 백테스트 (exit bar index 반환) ─────────────────────────────
@njit(cache=True)
def _backtest_idx(high, low, open_, close,
                  short_sig, long_sig,
                  rr_ratio, taker_fee, max_bars=1440):
    n = len(high)
    res_r   = np.empty(n, dtype=np.float64)
    res_idx = np.empty(n, dtype=np.int64)
    cnt = 0
    exit_idx = -1

    for i in range(n - 1):
        if i <= exit_idx:
            continue
        if not (short_sig[i] or long_sig[i]):
            continue

        if short_sig[i]:
            sl_price    = high[i]
            entry_price = open_[i + 1]
            sl_dist     = sl_price - entry_price
            side        = -1
        else:
            sl_price    = low[i]
            entry_price = open_[i + 1]
            sl_dist     = entry_price - sl_price
            side        = 1

        if sl_dist <= 0.0:
            continue

        tp_price = (entry_price - sl_dist * rr_ratio if side == -1
                    else entry_price + sl_dist * rr_ratio)

        lev       = entry_price / sl_dist
        fee_entry = taker_fee * lev
        fee_sl    = taker_fee * lev

        found = False
        j_end = min(i + max_bars + 1, n)
        for j in range(i + 1, j_end):
            if side == -1:
                hit_sl = high[j] >= sl_price
                hit_tp = low[j]  <= tp_price
            else:
                hit_sl = low[j]  <= sl_price
                hit_tp = high[j] >= tp_price

            if hit_sl and hit_tp:
                r = -1.0 - fee_entry - fee_sl
            elif hit_sl:
                r = -1.0 - fee_entry - fee_sl
            elif hit_tp:
                r = rr_ratio - fee_entry
            else:
                continue

            res_r[cnt]   = r
            res_idx[cnt] = j
            cnt += 1
            exit_idx = j
            found = True
            break

        if not found:
            j = min(i + max_bars, n - 1)
            r_raw = ((entry_price - close[j]) / sl_dist if side == -1
                     else (close[j] - entry_price) / sl_dist)
            res_r[cnt]   = r_raw - fee_entry - fee_sl
            res_idx[cnt] = j
            cnt += 1
            exit_idx = j

    return res_r[:cnt], res_idx[:cnt]


# ── 복리 월별 집계 ────────────────────────────────────────────────────
def compound_monthly(trades_df, risk_pct, init_cap=INIT_CAP):
    """trades_df: [연월(str), R(float)] → 월별 복리 수익% + 누적자본"""
    months = sorted(trades_df["연월"].unique())
    records = []
    capital = init_cap

    for ym in months:
        rs = trades_df[trades_df["연월"] == ym]["R"].values
        cap_start = capital
        trades_cnt = len(rs)
        for r in rs:
            capital *= (1.0 + risk_pct * r)
        mo_ret = (capital / cap_start - 1.0) * 100.0
        cum_ret = (capital / init_cap - 1.0) * 100.0
        records.append({
            "연월": ym,
            "거래수": trades_cnt,
            "월수익%": round(mo_ret, 2),
            "누적수익%": round(cum_ret, 2),
        })

    return pd.DataFrame(records)


# ── 트레이드 데이터 수집 ──────────────────────────────────────────────
def get_trades(df, params):
    short_sig, long_sig = detect_signals(
        df, params["big_mult"], params["cover_pct"], params["avg_len"])
    ts_arr = df["ts"].values  # numpy datetime64
    res_r, res_idx = _backtest_idx(
        df["high"].values, df["low"].values,
        df["open"].values, df["close"].values,
        short_sig, long_sig,
        params["rr_ratio"], params["taker_fee"])

    exit_ts = pd.DatetimeIndex(ts_arr[res_idx]).tz_localize("UTC")
    ym = exit_ts.strftime("%Y%m")
    return pd.DataFrame({"연월": ym, "R": res_r})


# ── 엑셀 헬퍼 ────────────────────────────────────────────────────────
def merge_title(ws, row, c1, c2, text, bg, fc=C_WHITE, sz=11):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font      = Font(name="맑은 고딕", bold=True, size=sz, color=fc)
    c.fill      = _fill(bg)
    c.alignment = _center()
    c.border    = _border()

def write_cell(ws, row, col, val, bold=False, sz=9, bg=None, fc="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="맑은 고딕", bold=bold, size=sz, color=fc)
    c.fill      = _fill(bg or C_WHITE)
    c.alignment = _center()
    c.border    = _border()
    return c


# ── 메인 ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== 복리 월별 성적 계산 시작 ===", flush=True)

    df_1m = load_1m()
    risk_labels = [f"{r*100:.1f}%" for r in RISK_PCT_LIST]

    wb = Workbook()
    wb.remove(wb.active)

    all_data = {}  # {tf: {구간: {risk_pct: monthly_df}}}

    for tf, params in PARAMS.items():
        print(f"\n[{tf}] 신호 감지 + 백테스트...", flush=True)
        df_tf = resample(df_1m, tf)

        df_is  = df_tf[(df_tf["ts"] >= IS_START)  & (df_tf["ts"] < IS_END)].reset_index(drop=True)
        df_oos = df_tf[(df_tf["ts"] >= OOS_START) & (df_tf["ts"] < OOS_END)].reset_index(drop=True)

        trades_is  = get_trades(df_is,  params)
        trades_oos = get_trades(df_oos, params)
        print(f"  IS 트레이드: {len(trades_is)}건 / OOS 트레이드: {len(trades_oos)}건", flush=True)

        all_data[tf] = {}
        for label, (구간, trades) in [("IS",  trades_is), ("OOS", trades_oos)]:
            all_data[tf][label] = {}
        for 구간, trades in [("IS", trades_is), ("OOS", trades_oos)]:
            for rp in RISK_PCT_LIST:
                mdf = compound_monthly(trades, rp)
                all_data[tf][구간][rp] = mdf

        # ── 시트 작성 ─────────────────────────────────────────────────
        ws = wb.create_sheet(f"{tf} 월별성적")
        ws.sheet_view.showGridLines = False

        N_RISK = len(RISK_PCT_LIST)  # 10
        # 컬럼 구성: A(여백) | B(연월) | C(거래수) | D~M(0.5%~5% 월수익%) | N(누적%, 마지막risk)
        # 실제: col B=2, C=3, D=4..M=13, N=14  → IS/OOS 블록 나란히
        # 심플하게: IS 블록 따로, OOS 블록 따로 (행 방향)

        COL = 2
        ROW = 2
        tf_col = TF_COLOR[tf]

        merge_title(ws, ROW, COL, COL + N_RISK + 1,
                    f"{tf} 타임프레임 — 복리 리스크별 월별 수익% (IS 2020~2022 / OOS 2023~2026-03)",
                    C_DARK, sz=12)
        ROW += 2

        param_str = (f"배수={params['big_mult']}  덮음={params['cover_pct']}  "
                     f"손익비={params['rr_ratio']}  평균봉={params['avg_len']}  "
                     f"수수료={params['taker_fee']*100:.3f}%")
        merge_title(ws, ROW, COL, COL + N_RISK + 1, param_str, tf_col, sz=10)
        ROW += 2

        for 구간, bg_hdr, bg_row in [("IS", C_IS_HDR, C_IS_BG), ("OOS", C_OOS_HDR, C_OOS_BG)]:
            period = "2020~2022" if 구간 == "IS" else "2023~2026-03"
            merge_title(ws, ROW, COL, COL + N_RISK + 1,
                        f"{구간}  ({period})", bg_hdr, sz=10)
            ROW += 1

            # 헤더 행
            write_cell(ws, ROW, COL,     "연월",  bold=True, sz=9, bg=bg_row, fc="000000")
            write_cell(ws, ROW, COL + 1, "거래수", bold=True, sz=9, bg=bg_row, fc="000000")
            for ci, lbl in enumerate(risk_labels):
                write_cell(ws, ROW, COL + 2 + ci, lbl, bold=True, sz=9, bg=bg_row, fc="000000")
            ROW += 1

            # 모든 월 수집
            all_months = set()
            for rp in RISK_PCT_LIST:
                mdf = all_data[tf][구간][rp]
                all_months.update(mdf["연월"].tolist())
            all_months = sorted(all_months)

            month_totals = {rp: 0.0 for rp in RISK_PCT_LIST}  # 누적 수익% 마지막값

            for mi, ym in enumerate(all_months):
                stripe = C_GRAY if mi % 2 == 1 else C_WHITE
                # 연월, 거래수 (첫 risk_pct 기준)
                mdf0 = all_data[tf][구간][RISK_PCT_LIST[0]]
                row0 = mdf0[mdf0["연월"] == ym]
                trades_cnt = int(row0["거래수"].iloc[0]) if len(row0) > 0 else 0

                write_cell(ws, ROW, COL,     ym,          sz=9, bg=stripe)
                write_cell(ws, ROW, COL + 1, trades_cnt,  sz=9, bg=stripe)

                for ci, rp in enumerate(RISK_PCT_LIST):
                    mdf = all_data[tf][구간][rp]
                    row_m = mdf[mdf["연월"] == ym]
                    val = round(row_m["월수익%"].iloc[0], 2) if len(row_m) > 0 else 0.0
                    bg = C_WIN if val >= 0 else C_LOSS
                    write_cell(ws, ROW, COL + 2 + ci, val, sz=9, bg=bg)
                    month_totals[rp] = (round(row_m["누적수익%"].iloc[0], 2)
                                        if len(row_m) > 0 else month_totals[rp])
                ROW += 1

            # 누적 수익% 행
            write_cell(ws, ROW, COL,     "누적수익%", bold=True, sz=9, bg=bg_row, fc="000000")
            write_cell(ws, ROW, COL + 1, "",           bold=True, sz=9, bg=bg_row)
            for ci, rp in enumerate(RISK_PCT_LIST):
                val = month_totals[rp]
                bg = C_WIN if val >= 0 else C_LOSS
                write_cell(ws, ROW, COL + 2 + ci, val, bold=True, sz=9, bg=bg)
            ROW += 2

        # 열 너비
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions[get_column_letter(COL)].width     = 8   # 연월
        ws.column_dimensions[get_column_letter(COL + 1)].width = 7   # 거래수
        for ci in range(N_RISK):
            ws.column_dimensions[get_column_letter(COL + 2 + ci)].width = 8
        ws.freeze_panes = f"{get_column_letter(COL)}5"

        print(f"  [{tf}] 시트 작성 완료", flush=True)

    output = "D:/candle-reversal/복리_리스크별_월별성적.xlsx"
    wb.save(output)
    print(f"\n저장완료: {output}", flush=True)
